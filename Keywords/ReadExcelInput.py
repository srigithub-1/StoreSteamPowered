# Import excel package
import openpyxl

#Load the data sheet
datasheet = openpyxl.load_workbook("C:\\Users\\USER\\PycharmProjects\\StoreSteamPowered\GameDetails.xlsx")

#Create object of the "Free to play" sheet
FTP = datasheet['Free to play']

#Extract the value of Row # 5 in "Free to play" sheet
#Option1:
#FTP['A5'].value

#Option2:
#cellobj = FTP.cell(2,1)
#print(cellobj.value)

#Option 3:
#cellobj = FTP.cell(column=1,row=2)
#print(cellobj.value)

#Find total number of rows
TotalRows = FTP.max_row
TotalColumns = FTP.max_column

#print("Total Number of Rows:"+str(TotalRows))
def List_PrintElements(games):
    print("First Element in the List:"+games[1])


